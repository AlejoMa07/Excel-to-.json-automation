import os
import time
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

# ==============================
# CONFIGURACIÓN DEL USUARIO
# ==============================

CARPETA_BASE = r"RUTA_AQUI"
URL_RIPS = "url"

# ==============================
# CONFIGURAR CHROME
# ==============================

carpeta_descargas = os.path.join(CARPETA_BASE, "temp_descargas")
os.makedirs(carpeta_descargas, exist_ok=True)

chrome_options = Options()
prefs = {
    "download.default_directory": carpeta_descargas,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True,

    # 🔥 PERMITIR DESCARGAS MÚLTIPLES
    "profile.default_content_setting_values.automatic_downloads": 1
}
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--disable-infobars")

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)
wait = WebDriverWait(driver, 30)

# Ventana pequeña (no molesta)
driver.set_window_size(900, 700)
driver.set_window_position(0, 0)

# ==============================
# FUNCIONES AUXILIARES
# ==============================

def obtener_numero_factura(archivo_excel):
    try:
        wb = load_workbook(archivo_excel, data_only=True)
        if "transaccion" not in wb.sheetnames:
            wb.close()
            return None
        hoja = wb["transaccion"]
        valor = hoja["B2"].value
        wb.close()
        return str(valor).strip() if valor else None
    except:
        return None


def esperar_descarga(carpeta, tiempo_max=60):
    for _ in range(tiempo_max):
        archivos = [f for f in os.listdir(carpeta) if f.lower().endswith(".json")]
        if archivos:
            return os.path.join(carpeta, archivos[0])
        time.sleep(1)
    return None


def cerrar_popup_forzado(driver):
    driver.execute_script("""
        const popups = document.querySelectorAll('.swal2-container');
        popups.forEach(p => p.remove());
        document.body.classList.remove('swal2-shown');
        document.body.style.overflow = 'auto';
    """)
    time.sleep(0.5)


def click_js(driver, element):
    driver.execute_script("arguments[0].click();", element)

# ==============================
# PROCESO PRINCIPAL
# ==============================

driver.get(URL_RIPS)
wait.until(EC.presence_of_element_located((By.ID, "inputGroupFileAddon04")))
print("🌐 Página cargada correctamente.")

for archivo in os.listdir(CARPETA_BASE):
    if archivo.lower().endswith((".xlsx", ".xlsm")) and "plantilla" in archivo.lower():
        ruta_excel = os.path.join(CARPETA_BASE, archivo)
        print(f"\n🟢 Procesando: {archivo}")

        num_factura = obtener_numero_factura(ruta_excel)
        if not num_factura:
            print("❌ No se encontró número de factura.")
            continue

        carpeta_destino = os.path.join(CARPETA_BASE, num_factura)
        os.makedirs(carpeta_destino, exist_ok=True)

        try:
            cerrar_popup_forzado(driver)

            # Subir archivo
            input_file = wait.until(
                EC.presence_of_element_located((By.ID, "inputGroupFileAddon04"))
            )
            input_file.send_keys(ruta_excel)

            # Convertir
            boton_convertir = wait.until(
                EC.presence_of_element_located((By.ID, "uploadButton"))
            )
            click_js(driver, boton_convertir)
            print("⏳ Convirtiendo...")

            # Confirmación
            try:
                boton_ok = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "button.swal2-confirm"))
                )
                click_js(driver, boton_ok)
            except:
                pass

            cerrar_popup_forzado(driver)
            print("✅ Confirmación gestionada.")

            # Descargar JSON
            archivo_json = esperar_descarga(carpeta_descargas, 60)
            if archivo_json:
                destino_json = os.path.join(
                    carpeta_destino, os.path.basename(archivo_json)
                )
                shutil.move(archivo_json, destino_json)
                print(f"✅ JSON guardado en {destino_json}")
            else:
                print("❌ No se detectó descarga.")

        except Exception as e:
            print(f"⚠️ Error procesando {archivo}: {e}")

        # Limpieza temporal
        for f in os.listdir(carpeta_descargas):
            try:
                os.remove(os.path.join(carpeta_descargas, f))
            except:
                pass

print("\n🏁 Proceso completado con éxito.")
driver.quit()
